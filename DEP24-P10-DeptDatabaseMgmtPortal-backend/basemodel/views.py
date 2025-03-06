from rest_framework.response import Response
from rest_framework import status 
from rest_framework.views import APIView
from .models import BaseModel
from .services import *
from .access import *
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated
import csv
from usercustom.models import CustomUser
from department.models import Department
import openpyxl
from django.http import HttpResponse
import pandas as pd
import io
from django.db import transaction
from django.contrib.auth.models import Group
from datetime import datetime
from logger_config import logger

class baseModelUserList(APIView):
    """
        Base class for all user list views
        default route is /user/
        default methods are get
        JWT auth is used
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseServices()
    model = BaseModel

    def get(self,request):
        if request.user.groups.filter(name = 'Staff').exists():
            kwargs = {}
        else:
            self.access.set_user(request.user)
            access,kwargs = self.access.can_get_user_list()
            if not access:
                return Response({'data':None,'errors':'You are not authorized to view '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.get_user_list(**kwargs)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)

class baseModelCreate(APIView):
    """
        Base class for all create views
        default route is /create/
        default methods are post
        JWT auth is used
    """
    authentication_classes = [JWTAuthentication]
    # permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseServices()
    model = BaseModel

    def post(self,request):
        self.access.set_user(request.user)
        if not self.access.can_create():

            return Response({'data':None,'errors':'You are not authorized to create '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.create(request)
        if stat:

            user = CustomUser.objects.get(id = request.user.id)
            print(f"Created by ID {user.email} Above Submitted with Resource ID {serializer.data['id']}")
            logger.info(f"Created by ID {user.email} with Submitted Resource ID {serializer.data['id']}")
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)

class baseModelDetailView(APIView):
    """
        Base class for all detail views
        default route is /<int:id>/
        default methods are get(view), put(update), delete(delete)
        JWT auth is used
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseServices()
    model = BaseModel

    def get(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_view(id):
            return Response({'data':None,'errors':'You are not authorized to view '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.get(id)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)
    
    def put(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_update(id):
            return Response({'data':None,'errors':'You are not authorized to update '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.update(id,request)
        if not stat:
            return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)
        return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
    
    def delete(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_delete(id):
            print("helo")
            return Response({'data':None,'errors':'You are not authorized to delete '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.delete(id,request)
        if not stat:
            return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)
        return Response({'data':serializer,'errors':None},status = status.HTTP_200_OK)  

class baseModelRestoreView(APIView):
    """
        Base class for all restore views
        default route is /restore/<int:id>/
        default methods are put(restore),get(view deleted object)
        JWT auth is used
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseServices()
    model = BaseModel

    def put(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_restore(id):
            return Response({'data':None,'errors':'You are not authorized to restore '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.restore(id,request)
        if not stat:
            return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)
        return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
    
    def get(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_view_deleted(id):
            return Response({'data':None,'errors':'You are not authorized to view '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.get_deleted_obj(id)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)

class baseModelDraftListView(APIView):
    """
        Base class for all draft list views
        default route is /draft/
        default methods are get(view), post(create)
        JWT auth is used
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseDraftServices()
    model = BaseModel

    def get(self,request):
        self.access.set_user(request.user)
        access,kwargs = self.access.can_view_draft_list()
        if not access:
            return Response({'data':None,'errors':'You are not authorized to view draft list '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.get_drafts(request,**kwargs)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)

    def post(self,request):
        self.access.set_user(request.user)
        if not self.access.can_create_draft():
            return Response({'data':None,'errors':'You are not authorized to create draft '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.create_draft(request)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)

class baseModelDraftView(APIView):
    """
        Base class for all draft views
        default route is /draft/<int:id>/
        default methods are get(view), put(update), delete(delete),post(create draft for existing object)
        JWT auth is used
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseDraftServices()
    model = BaseModel

    def post(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_create_previous_draft(id):
            return Response({'data':None,'errors':'You are not authorized to create draft '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.create_previous_draft(id, request)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)

    def get(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_view_draft(id):
            return Response({'data':None,'errors':'You are not authorized to view draft '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.get_draft(id)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)
    
    def put(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_update_draft(id):
            return Response({'data':None,'errors':'You are not authorized to update draft '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.update_draft(id, request)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)
    
    def delete(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_delete_draft(id):
            return Response({'data':None,'errors':'You are not authorized to delete draft '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.delete_draft(id, request)
        if stat:
            return Response({'data':serializer,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)
     
class baseModelSubmitView(APIView):
    """
        Base class for all submit views
        default route is /submit/<int:id>/
        default methods are put(submit)
        JWT auth is used
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseDraftServices()
    model = BaseModel

    def put(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_submit_draft(id):
            return Response({'data':None,'errors':'You are not authorized to submit draft '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.submit(id, request)
        if stat:
            user = CustomUser.objects.get(id = request.user.id)
            print(f"Submission by ID {user.email} Above Submitted with Resource ID {serializer.data['id']}")
            logger.info(f"Submission by ID {user.email} with Submitted Resource ID {serializer.data['id']}")
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)

class baseModelApproveView(APIView):
    """
        Base class for all approve views
        default route is /approve/<int:id>/
        default methods are get(view), put(approve)
        JWT auth is used
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseDraftServices()
    model = BaseModel

    def get(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_view_pending(id):
            return Response({'data':None,'errors':'You are not authorized to view pending '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.get_pending_obj(id)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)

    def put(self,request,id):
        self.access.set_user(request.user)
        print("Hey this Put",self)
        if not self.access.can_approve(id):
            return Response({'data':None,'errors':'You are not authorized to approve draft '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.approve(id, request)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)

class baseModelRejectView(APIView):
    """
        Base class for all reject views
        default route is /reject/<int:id>/
        default methods are put(reject)
        JWT auth is used
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseDraftServices()
    model = BaseModel

    def put(self,request,id):
        self.access.set_user(request.user)
        if not self.access.can_reject(id):
            return Response({'data':None,'errors':'You are not authorized to reject draft '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.reject(id, request)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)

class baseModelPendingListView(APIView):
    """
        Base class for all pending list views
        default route is /pending/
        default methods are get(view)
        JWT auth is used
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseDraftServices()
    model = BaseModel

    def get(self,request):
        self.access.set_user(request.user)
        access,kwargs1,kwargs2 = self.access.can_view_pending_list()
        if not access:
            return Response({'data':None,'errors':'You are not authorized to view pending list '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.get_pending(request,kwargs1,kwargs2)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)

class baseModelDeletedListView(APIView):
    """    
        Base class for all deleted list views
        default route is /deleted/
        default methods are get(view)
        JWT auth is used
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseServices()
    model = BaseModel

    def get(self,request):
        self.access.set_user(request.user)
        access,kwargs = self.access.can_view_deleted_list()
        if not access:
            return Response({'data':None,'errors':'You are not authorized to view deleted list '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)
        stat, serializer = self.service.get_deleted(request,**kwargs)
        if stat:
            return Response({'data':serializer.data,'errors':None},status = status.HTTP_200_OK)
        return Response({'data':None,'errors':serializer},status = status.HTTP_400_BAD_REQUEST)
    

class BaseListUploadView(APIView):
    authentication_classes = [JWTAuthentication]
    # permission_classes = [IsAuthenticated]
    access = BaseAccessSpecifier()
    service = BaseServices()
    model = BaseModel
    serializer = BaseSerializer
    req = None
    filename = None
    return_filename = None

    def add_validators(self, ws):
        for i in range(1000):
            rule = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1 = 'sheet2!$A2:$A99999', allow_blank=True)
            rule.error = 'Please select a valid option'
            rule.errorTitle = 'Invalid option'
            ws.add_data_validation(rule)
            rule.add(f'D{i+3}')
        for i in range(1000):
            rule = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1 = 'sheet3!$A2:$A99999', allow_blank=True)
            rule.error = 'Please select a valid option'
            rule.errorTitle = 'Invalid option'
            ws.add_data_validation(rule)
            rule.add(f'F{i+3}')
        for i in range(1000):
            rule = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1 = 'sheet4!$A2:$A99999', allow_blank=True)
            rule.error = 'Please select a valid option'
            rule.errorTitle = 'Invalid option'
            ws.add_data_validation(rule)
            rule.add(f'H{i+3}')
        return ws  
        

    def get(self, request):
        try:
            student_group = Group.objects.get(name='Student')
            faculty_group = Group.objects.get(name='Faculty')
            students = CustomUser.objects.filter(groups = student_group)
            faculty = CustomUser.objects.filter(groups = faculty_group)
            users = (students | faculty).distinct()
            departments = Department.objects.all()
            data = [[obj.first_name + ' ' + obj.last_name, obj.email] for obj in users]
            data = [['Name', 'Email'], *data]
            dept_data = [[obj.code, obj.id] for obj in departments]
            dept_data = [['Code', 'ID'], *dept_data]
            wb = openpyxl.load_workbook(self.filename,read_only=False, keep_vba=True)
            ws = wb['Sheet3']
            ws.delete_rows(1, ws.max_row)
            for row in data:
                ws.append(row)
            ws = wb['Sheet4']
            ws.delete_rows(1, ws.max_row)
            for row in dept_data:
                ws.append(row)
            ws = wb['Sheet1']
            ws = self.add_validators(ws)
            response = HttpResponse(content = openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            fime_name_return = self.return_filename[:-5] +str(datetime.now())  +'.xlsm'
            response['Content-Disposition'] = f'inline; filename="{fime_name_return}"'
            return response
        except Exception as e:
            print(e)
            return Response({'error':str(e)},status=status.HTTP_400_BAD_REQUEST)
        
    def post(self, request, *args, **kwargs):
        self.req = request
        self.access.set_user(request.user)
        print(request.FILES)
        print(request.FILES['file'])
        if self.access.can_create() or self.access.is_staff():
            print('start')
            try:
                flag = False
                print(request.FILES['file'])
                csv_file = request.FILES['file']
                if not csv_file.name.endswith('.csv'):
                    print("Not csv")
                    return Response({'error':'File is not CSV type'},status=status.HTTP_400_BAD_REQUEST)
                else:
                    decoded_file = csv_file.read().decode('utf-8-sig').splitlines()
                    print(decoded_file)
                    csvFile = csv.DictReader(decoded_file)
                    datatypes = next(csvFile)
                count = 1
                print("parsing")
                values = {}
                error_values = {}
                # if()

                with transaction.atomic():
                    for row in csvFile:
                        try:
                            print("calling function")
                            values[count], error_values[count] = add_model(row, datatypes, request, self.serializer)
                            print("success")
                            if error_values[count] != None:
                                # specific_error = {'department': [ErrorDetail(string='This field may not be null.', code='null')], 'title': [ErrorDetail(string='This field may not be null.', code='null')], 'publication_type': [ErrorDetail(string='This field may not be null.', code='null')], 'publication_status': [ErrorDetail(string='This field may not be null.', code='null')], 'identifier_type': [ErrorDetail(string='This field may not be null.', code='null')], 'authors': [ErrorDetail(string='This field may not be null.', code='null')]}
                                specific_error = {'department': ['This field may not be null.'], 'title': ['This field may not be null.'], 'publication_type': ['This field may not be null.'], 'publication_status': ['This field may not be null.'], 'identifier_type': ['This field may not be null.'], 'authors': ['This field may not be null.']}

                                if error_values[count] != specific_error:
                                    flag = True
                                    raise Exception(error_values[count])
                        except Exception as e:
                            print("exception error: ",str(e))
                            error_values[count] = str(e)
                            values[count] = None
                            flag = True
                            raise Exception(str(e))
                        count += 1
                    print("success")
                    return Response({'data':values,'data_errors':error_values,'errors':None},status = status.HTTP_200_OK)
                # with transaction.atomic():
                #     for row in csvFile:
                #         # try:
                #             print("calling function")
                #             values[count], error_values[count] = add_model(row, datatypes, request, self.serializer)
                #             print("success")
                #             # if error_values[count] != None:
                #                 # flag = True
                #                 # raise Exception(error_values[count])
                #         # except Exception as e:
                #         #     print("exception error: ",str(e))
                #         #     error_values[count] = str(e)
                #         #     values[count] = None
                #         #     flag = True
                #         #     raise Exception(str(e))
                #         # count += 1
                #     print("success")
                #     return Response({'data':values,'data_errors':error_values,'errors':None},status = status.HTTP_200_OK)
            except Exception as e:
                print("error")
                print(e)
                if flag:
                    return Response({'data':values,'data_errors':error_values,'errors':str(e)},status = status.HTTP_400_BAD_REQUEST)
                return Response({'errors':str(e)},status=status.HTTP_400_BAD_REQUEST)
        else:
            print('You are not authorized to create '+self.model.__name__)
            return Response({'errors':'You are not authorized to create '+self.model.__name__},status = status.HTTP_401_UNAUTHORIZED)